import os
import pandas as pd
from dataclasses import dataclass, asdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from typing import List
from datetime import datetime

from utils.shared import JobListing


class JobsExcelExporter:
    def __init__(self, excel_path: str):
        """
        Initialize the Excel exporter with the path where the Excel file should be saved
        
        Args:
            excel_path (str): Full path including filename where the Excel should be saved
        """
        self.excel_path = self._get_valid_path(excel_path)

    def _get_valid_path(self, path: str) -> str:
        """
        Ensure the file path is accessible; if not, append the current date to the filename.

        Args:
            path (str): The original file path.

        Returns:
            str: A valid file path.
        """
        directory, filename = os.path.split(path)
        if not os.path.exists(directory) and directory:
            print(f"Warning: Path '{directory}' is inaccessible. Using fallback filename.")
            name, ext = os.path.splitext(filename)
            new_filename = f"{name}_{datetime.now().strftime('%Y%m%d')}{ext}"
            return os.path.join(os.getcwd(), new_filename)  # Save in the current directory
        return path

    def _format_column_title(self, title: str) -> str:
        """
        Format column titles to be more readable
        
        Args:
            title (str): Original column title
            
        Returns:
            str: Formatted column title
        """
        words = title.replace('_', ' ').split()
        return ' '.join(word.capitalize() for word in words)

    def get_column_letter(self, n: int):
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string
    
    def export_jobs(self, jobs: List[JobListing], sheet_name: str = "LinkedIn Jobs"):
        """
        Export jobs to Excel and format them as a table
        
        Args:
            jobs (List[LinkedinJobListing]): List of LinkedIn job listings to export
            sheet_name (str): Name of the sheet where jobs should be exported
        """
        # Convert jobs to pandas DataFrame with special handling for List fields
        jobs_data = []
        for job in jobs:
            job_dict = vars(job)
            # Convert List[str] to comma-separated string
            job_dict['technologies_required'] = ', '.join(
                job_dict['technologies_required'])
            
            jobs_data.append(job_dict)

        df = pd.DataFrame(jobs_data)

        # Define the column order based on the dataclass fields
        column_order = [
            'title', 'date_analyzed', 'posted_time',
            'experience_years_needed', 'easy_apply',
            'seniority_level', 'employment_type',
            'job_function', 'industries',
            'required_studies', 'technologies_required',
            'company', 'location', 'salary_offered',
            'job_id', 'title_lang', 'description_lang',
            'source', 'url'  # URL always last
        ]

        # Ensure all columns exist in the DataFrame
        for col in column_order:
            if col not in df.columns:
                df[col] = ''  # Add empty column if missing

        # Reorder columns
        df = df[column_order]
        
        df = df.sort_values(by=['date_analyzed', 'posted_time'], ascending=[False, False])

        # Format column names
        formatted_columns = {
            col: self._format_column_title(col) for col in df.columns}
        df = df.rename(columns=formatted_columns)

        # Export to Excel
        df.to_excel(self.excel_path, sheet_name=sheet_name, index=False)

        # Load workbook for formatting
        wb = load_workbook(self.excel_path)
        ws = wb[sheet_name]
        
        # Define a date style
        date_style = NamedStyle(name="datetime", number_format="DD of MMMM YYYY")
        if "datetime" not in wb.named_styles:
            wb.add_named_style(date_style)

        # Apply date style to 'date_analyzed' and 'posted_time' columns
        date_analyzed_col_letter = self.get_column_letter(df.columns.get_loc('Date Analyzed') + 1)
        posted_time_col_letter = self.get_column_letter(df.columns.get_loc('Posted Time') + 1)

        for row in range(2, len(jobs) + 2):  # Start from row 2 to skip header
            ws[f'{date_analyzed_col_letter}{row}'].style = 'datetime'
            ws[f'{posted_time_col_letter}{row}'].style = 'datetime'

        last_column_letter = self.get_column_letter(len(df.columns))
        table_ref = f"A1:{last_column_letter}{len(jobs) + 1}"
        table = Table(displayName="JobListings", ref=table_ref)

        # Add a style to the table
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # Format headers
        header_fill = PatternFill(start_color="366092",
                                end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)

        # Format URL column
        url_col_letter = self.get_column_letter(
            len(df.columns))  # Changed to use the function
        for row in range(2, len(jobs) + 2):
            cell = ws[f'{url_col_letter}{row}']
            if cell.value:  # If there's a URL
                cell.hyperlink = cell.value
                cell.value = "Link"
                # Blue, underlined
                cell.font = Font(color="0563C1", underline="single")
                cell.alignment = Alignment(horizontal='center')

        # Apply conditional formatting
        self._apply_conditional_formatting(ws, len(jobs))

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            # Make URL column narrower since it just contains "Link"
            if column[0].column_letter == url_col_letter:
                ws.column_dimensions[column[0].column_letter].width = 10
            else:
                ws.column_dimensions[column[0].column_letter].width = min(
                    adjusted_width, 50)

        # Save the workbook
        wb.save(self.excel_path)
    
    def _apply_conditional_formatting(self, worksheet, row_count: int):
        """
        Apply conditional formatting to specific columns
        
        Args:
            worksheet: The worksheet to format
            row_count (int): Number of rows with data
        """
        # Example: Highlight non-English listings in yellow
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Find title_lang and desc_lang column indices
        header_row = list(worksheet[1])
        title_lang_idx = None
        desc_lang_idx = None
        
        for idx, cell in enumerate(header_row, 1):
            if cell.value == "Title Lang":
                title_lang_idx = idx
            elif cell.value == "Description Lang":
                desc_lang_idx = idx
        
        if title_lang_idx:
            for row in range(2, row_count + 2):
                cell = worksheet.cell(row=row, column=title_lang_idx)
                if cell.value and cell.value.lower() != "en":
                    cell.fill = yellow_fill
        
        if desc_lang_idx:
            for row in range(2, row_count + 2):
                cell = worksheet.cell(row=row, column=desc_lang_idx)
                if cell.value and cell.value.lower() != "en":
                    cell.fill = yellow_fill
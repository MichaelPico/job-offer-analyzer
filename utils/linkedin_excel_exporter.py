import os
import pandas as pd
from dataclasses import dataclass, asdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from typing import List

from utils.shared import LinkedinJobListing


class LinkedinExcelExporter:
    def __init__(self, excel_path: str):
        """
        Initialize the Excel exporter with the path where the Excel file should be saved
        
        Args:
            excel_path (str): Full path including filename where the Excel should be saved
        """
        self.excel_path = excel_path
        
    def _format_column_title(self, title: str) -> str:
        """
        Format column titles to be more readable
        
        Args:
            title (str): Original column title
            
        Returns:
            str: Formatted column title
        """
        # Split by underscore and capitalize each word
        words = title.replace('_', ' ').split()
        return ' '.join(word.capitalize() for word in words)
    
    def export_jobs(self, jobs: List[LinkedinJobListing], sheet_name: str = "LinkedIn Jobs"):
        """
        Export jobs to Excel and format them as a table
        
        Args:
            jobs (List[LinkedinJobListing]): List of LinkedIn job listings to export
            sheet_name (str): Name of the sheet where jobs should be exported
        """
        # Convert jobs to pandas DataFrame
        jobs_data = ([vars(job) for job in jobs])
        df = pd.DataFrame(jobs_data)
        
        # Reorder columns to put URL at the end
        columns = [col for col in df.columns if col != 'url'] + ['url']
        df = df[columns]
        
        # Create formatted column names dictionary
        formatted_columns = {col: self._format_column_title(col) for col in df.columns}
        df = df.rename(columns=formatted_columns)
        
        # Export to Excel
        df.to_excel(self.excel_path, sheet_name=sheet_name, index=False)
        
        # Load workbook for formatting
        wb = load_workbook(self.excel_path)
        ws = wb[sheet_name]
        
        # Format as table
        table_ref = f"A1:{chr(65 + len(df.columns) - 1)}{len(jobs) + 1}"
        table = Table(displayName="JobListings", ref=table_ref)
        
        # Add a style to the table
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Format headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Format URL column
        url_col_letter = chr(65 + len(df.columns) - 1)
        for row in range(2, len(jobs) + 2):
            cell = ws[f'{url_col_letter}{row}']
            if cell.value:  # If there's a URL
                cell.hyperlink = cell.value
                cell.value = "Link"
                cell.font = Font(color="0563C1", underline="single")  # Blue, underlined
                cell.alignment = Alignment(horizontal='center')
        
        # Apply conditional formatting
        self._apply_conditional_formatting(ws, len(jobs))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            # Make URL column narrower since it just contains "Link"
            if column[0].column_letter == url_col_letter:
                ws.column_dimensions[column[0].column_letter].width = 10
            else:
                ws.column_dimensions[column[0].column_letter].width = min(adjusted_width, 50)
        
        # Save the workbook
        wb.save(self.excel_path)
    
    def _apply_conditional_formatting(self, worksheet, row_count: int):
        """
        Apply conditional formatting to specific columns
        
        Args:
            worksheet: The worksheet to format
            row_count (int): Number of rows with data
        """
        # Example: Highlight non-English listings in yellow
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Find title_lang and desc_lang column indices
        header_row = list(worksheet[1])
        title_lang_idx = None
        desc_lang_idx = None
        
        for idx, cell in enumerate(header_row, 1):
            if cell.value == "Title Lang":
                title_lang_idx = idx
            elif cell.value == "Description Lang":
                desc_lang_idx = idx
        
        if title_lang_idx:
            for row in range(2, row_count + 2):
                cell = worksheet.cell(row=row, column=title_lang_idx)
                if cell.value and cell.value.lower() != "en":
                    cell.fill = yellow_fill
        
        if desc_lang_idx:
            for row in range(2, row_count + 2):
                cell = worksheet.cell(row=row, column=desc_lang_idx)
                if cell.value and cell.value.lower() != "en":
                    cell.fill = yellow_fill